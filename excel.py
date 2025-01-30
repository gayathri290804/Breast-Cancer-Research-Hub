import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import os

def upload_excel():
    # Define the relative path to your Excel file
    file_path = "Final Data.xlsx"  # Using relative path

    if not os.path.exists(file_path):
        st.error(f"Error: The file '{file_path}' was not found. Please check if it is uploaded to the repository.")
        return

    try:
        # Load the workbook
        wb = load_workbook(file_path, data_only=True)
        sheet = wb.active

        # Extract data and hyperlinks
        data = []
        hyperlinks = []
        for row in sheet.iter_rows():
            row_data = []
            row_links = []
            for cell in row:
                row_data.append(cell.value)
                row_links.append(cell.hyperlink.target if cell.hyperlink else None)
            data.append(row_data)
            hyperlinks.append(row_links)

        # Convert data and links into DataFrames
        headers = data[0]
        df = pd.DataFrame(data[1:], columns=headers)
        links_df = pd.DataFrame(hyperlinks[1:], columns=headers)

        # Add Serial Number (S No.) as a new column
        df.insert(0, 'S No.', range(1, len(df) + 1))

        # List of columns to process for hyperlinks
        link_columns = ['3D Structure', 'Sequence', 'PubMed ID', 'DOI ID', 'Uniprot', 'BLAST', 'Conserved Domain']

        # Add clickable labels for links in the DataFrame
        for col in link_columns:
            if col in df.columns and col in links_df.columns:
                df[col] = df.apply(
                    lambda x: f'<a href="{links_df.at[x.name, col]}" target="_blank">{x[col]}</a>'
                    if pd.notna(links_df.at[x.name, col]) else x[col],
                    axis=1,
                )

        # Search functionality
        search_query = st.text_input("Search by Name or Keyword", value="")
        if search_query:
            # Filter DataFrame based on search query
            filtered_df = df[df.apply(lambda row: row.astype(str).str.contains(search_query, case=False).any(), axis=1)]
            if not filtered_df.empty:
                # Highlight rows matching the query
                highlight_style = "background-color: grey; font-weight: bold;"
                df_styles = filtered_df.style.applymap(
                    lambda val: highlight_style if search_query.lower() in str(val).lower() else "",
                )
                # Display the filtered and highlighted DataFrame
                st.markdown(df_styles.to_html(escape=False), unsafe_allow_html=True)
            else:
                st.warning("No results found for your search.")
        
        else:
            # Convert the DataFrame to an HTML table
            st.markdown(df.to_html(escape=False, index=False), unsafe_allow_html=True)

        # Apply CSS styles
        st.markdown("""
            <style>
                .search-bar {
                    margin-bottom: 10px;
                }
                .search-bar input {
                    border: 1px solid black;
                    border-radius: 5px;
                    box-shadow: 2px 2px 5px rgba(0, 0, 0, 0.1);
                }
                table {
                    width: 50%;
                    border-collapse: collapse;
                }
                th, td {
                    border: 1px solid #ddd;
                    padding: 8px;
                    text-align: left;
                }
                tr:nth-child(even) {
                    background-color: #f2f2f2;
                }
                tr:hover {
                    background-color: #ddd;
                }
                th {
                    background-color: #4CAF50;
                    color: white;
                    text-align: center;
                }
            </style>
        """, unsafe_allow_html=True)

    except Exception as e:
        st.error(f"An error occurred while loading the Excel file: {e}")
